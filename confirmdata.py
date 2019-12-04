"""导入核心期刊全部文章和引证文章，删除引证文章中不属于核心期刊的引用"""
"""涉及python操作Excel和一些列表，循环操作"""
import openpyxl
import re
# 获取文献库对象
workbook = openpyxl.load_workbook("confirmdata.xlsx")
# 获取工作簿 workbook的所有工作表
shenames = workbook.sheetnames
print(shenames)
#获取工作对象
worksheet=workbook[shenames[0]]
#print(worksheet)
#获取该表的行数和列数
rows=worksheet.max_row
columns=worksheet.max_column
print(rows,columns)
confirmdata = []


#按行获取数据
for row in worksheet.rows:
    for cell in row:
        print(cell.value,end=" ")
        confirmdata.append(cell.value)
    #print()
print(confirmdata[1])

confirmdata1 =[]
for i in confirmdata:
    if i is not None:
        confirmdata1.append(i)
    else:
        continue

workbook = openpyxl.load_workbook("oridata.xlsx")
# 获取工作簿 workbook的所有工作表
shenames = workbook.sheetnames
print(shenames)
#获取工作对象
worksheet=workbook[shenames[0]]
print(worksheet)
#获取该表的行数和列数
rows=worksheet.max_row
columns=worksheet.max_column
print(rows,columns)
data = []
#按行获取数据
for row in worksheet.rows:
    for cell in row:
        #print(cell.value,end=" ")
        data.append(cell.value)
    #print()

#分割列表中的文献
re_result1 = []
for i in confirmdata1:
    re_result = re.split('\n',i)
    re_result1.append(re_result)
#print(re_result1)
print(re_result1)  #复杂分割,不同的分隔符中间用“\n”进行间隔

#判断分割后的文献是否在文献库中，不在的删除
# import xlwt
# xls = xlwt.Workbook()
# # 设置字体格式
# Font0 = xlwt.Font()
# Font0.name = "Times New Roman"
# Font0.colour_index = 2
# style0 = xlwt.XFStyle()
# # 实例化一个工作表，名叫Sheet1
# sht1 = xls.add_sheet('Sheet1')

finaldata = [ [0] * 30 for i in range(1500)]
#print(finaldata)
a = 0
b = 0
for i in re_result1:
    for p in i:
        for j in data:
            if j in p:
                finaldata[a][b] = p
                b = b+1
            else:
                continue
    a = a+1
    b = 0


#按索引删除二维数组中的0
# for p in range(len(finaldata)):
#     for q in range(len(finaldata[p])):
#         if finaldata[p][q] == 0:
#             del finaldata[p][q]
#             q = q-1
#         else:
#             continue
# for i in finaldata:
#     for p in i:
#         if p == 0:
#             delete = finaldata.pop(p)
#         else:
#             continue
# print(finaldata[0])

# coding:utf-8
import xlsxwriter

workbook = xlsxwriter.Workbook('demo1.xlsx')  # 创建一个excel文件
worksheet= workbook.add_worksheet(u'sheet1')#在文件中创建一个名为TEST的sheet,不加名字默认为sheet1
# worksheet.set_column('A:A',20)#设置第一列宽度为20像素
bold= workbook.add_format({'bold':True})#设置一个加粗的格式对象
p = q = 0
count = 0
countdata = []
for i in finaldata:
    for j in i:
        if j != 0:
            count += 1
        else:
            continue
    countdata.append(count)
    count =0

for i in finaldata:
    for j in i:
        worksheet.write(p, q, j)
        #将j的元素按一行写入
        q = q+1
    p = p+1
    q = 0
workbook.close()
print(countdata)